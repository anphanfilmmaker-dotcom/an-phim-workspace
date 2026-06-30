import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from db_connection import execute_query
import openpyxl

if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

def import_excel():
    wb = openpyxl.load_workbook('g:/My Drive/[ANPHIM] MASTER PLANN/Project Management .xlsx', data_only=True)
    sheet_proj = wb['Projects']
    sheet_docs = wb['Giấy tờ']
    
    rows_proj = list(sheet_proj.iter_rows(values_only=True))[1:] # Skip header
    rows_docs = list(sheet_docs.iter_rows(values_only=True))[1:] # Skip header
    
    for i in range(len(rows_proj)):
        p_row = rows_proj[i]
        d_row = rows_docs[i] if i < len(rows_docs) else None
        
        if not p_row or not p_row[4]: # No project name
            continue
            
        proj_id_raw = str(p_row[0]).strip()
        proj_id = f"PROJ_{proj_id_raw}"
        client_raw = str(p_row[1]).strip() if p_row[1] else None
        proj_name = str(p_row[4]).strip()
        
        # Determine client_id
        client_id = None
        if client_raw:
            client_id = "VISTA_X" if client_raw.lower() == "vistax" else client_raw.upper()
            
        status = str(p_row[8]).strip() if p_row[8] else None
        budget = int(d_row[10]) if d_row and d_row[10] else 0
        
        print(f"Importing {proj_name}...")
        
        # 1. Insert into projects
        q_proj = """
            INSERT INTO projects (id, name, client_id, status, budget)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE SET
                name = EXCLUDED.name,
                client_id = EXCLUDED.client_id,
                status = EXCLUDED.status,
                budget = EXCLUDED.budget
        """
        execute_query(q_proj, (proj_id, proj_name, client_id, status, budget), fetch=False)
        
        # 2. Insert into projectDocuments
        if d_row:
            overall = str(d_row[2]).strip() if d_row[2] else ""
            quote = True if str(d_row[3]).strip().lower() == 'x' else False
            contract = True if str(d_row[4]).strip().lower() == 'x' else False
            vatR1 = True if str(d_row[5]).strip().lower() == 'x' else False
            vatR2 = True if str(d_row[6]).strip().lower() == 'x' else False
            vatR3 = True if str(d_row[7]).strip().lower() == 'x' else False
            liquidation = True if str(d_row[8]).strip().lower() == 'x' else False
            
            q_doc = """
                INSERT INTO projectDocuments (projectId, projectName, overallStatus, quote, contract, vatR1, vatR2, vatR3, liquidation)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (projectId) DO UPDATE SET
                    projectName = EXCLUDED.projectName,
                    overallStatus = EXCLUDED.overallStatus,
                    quote = EXCLUDED.quote,
                    contract = EXCLUDED.contract,
                    vatR1 = EXCLUDED.vatR1,
                    vatR2 = EXCLUDED.vatR2,
                    vatR3 = EXCLUDED.vatR3,
                    liquidation = EXCLUDED.liquidation
            """
            execute_query(q_doc, (proj_id, proj_name, overall, quote, contract, vatR1, vatR2, vatR3, liquidation), fetch=False)
            
    print("Excel import completed!")

if __name__ == "__main__":
    import_excel()
