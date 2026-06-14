import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'g:\My Drive\[ANPHIM] MASTER PLANN\Project Management .xlsx', data_only=True)
print("Sheets:", wb.sheetnames)
if 'Giấy tờ' in wb.sheetnames:
    sheet = wb['Giấy tờ']
    print(list(sheet.iter_rows(min_row=1, max_row=2, values_only=True)))
else:
    print("No Giấy tờ sheet")
